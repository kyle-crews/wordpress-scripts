import openpyxl

def generate_html_for_each_tab(excel_file_path):
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet_names = workbook.sheetnames

    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        filename = f"{sheet_name.replace('/', '-').replace('//', '-')}.txt"  # Replace forbidden characters in filenames
        with open(filename, 'w', encoding='utf-8') as output_file:
            html_str = "<div class='tablehead'>\n"
            html_str += f"  <h3>{sheet_name}</h3>\n"
            html_str += "  <button class='theme-button'>\n"
            html_str += "    VIEW ALL PAST CHAMPIONS\n"
            html_str += "  </button>\n"
            html_str += "</div>\n"
            html_str += "<table>\n"

            # Processing headers and rows here...

            headers = [cell.value for cell in sheet[1] if cell.value]
            if headers:
                html_str += "<tr>\n"
                for header in headers:
                    html_str += f"  <th class='fontbold'>{header}</th>\n"
                html_str += "</tr>\n"

            for row in sheet.iter_rows(min_row=2, values_only=False):
                row_values = [cell.value for cell in row]
                if not any(row_values):
                    continue

                html_str += "<tr class='tablerow'>\n"
                for i, cell in enumerate(row):
                    if i >= len(headers):
                        break

                    cell_value = cell.value if cell.value else ''
                    if i == 0 and isinstance(cell_value, float):
                        cell_value = int(cell_value)

                    if cell.hyperlink:
                        parts = str(cell_value).rsplit(' ', 1)
                        link_text = parts[-1] if len(parts) > 1 else cell_value
                        non_link_text = parts[0] if len(parts) > 1 else ''
                        cell_value = f'{non_link_text} <a href="{cell.hyperlink.target}">{link_text}</a>'
                    else:
                        cell_value = str(cell_value)

                    html_str += f"  <td>{cell_value}</td>\n"
                html_str += "</tr>\n"
            html_str += "</table>\n\n"

            output_file.write(html_str)

# Example usage
generate_html_for_each_tab('./TGA_Events_Past_Champions.xlsx')




# import pandas as pd
# import openpyxl

# def generate_html_from_excel_with_adjustments(excel_file_path, output_file_path):
#     workbook = openpyxl.load_workbook(excel_file_path)
#     sheet_names = workbook.sheetnames

#     with open(output_file_path, 'w') as output_file:
#         for sheet_name in sheet_names:
#             sheet = workbook[sheet_name]
#             html_str = "<div class='tablehead'>\n"
#             html_str += f"  <h3>{sheet_name}</h3>\n"
#             html_str += "  <button class='theme-button'>\n"
#             html_str += "    VIEW ALL PAST CHAMPIONS\n"
#             html_str += "  </button>\n"
#             html_str += "</div>\n"
#             html_str += "<table>\n"

#             # Process header row, skipping empty headers
#             headers = [cell.value for cell in sheet[1] if cell.value]
#             if headers:
#                 html_str += "<tr>\n"
#                 for header in headers:
#                     html_str += f"  <th class='fontbold'>{header}</th>\n"
#                 html_str += "</tr>\n"

#             # Process data rows
#             for row in sheet.iter_rows(min_row=2, values_only=False):
#                 row_values = [cell.value for cell in row]
#                 # Skip rows where all cells are empty
#                 if not any(row_values):
#                     continue

#                 html_str += "<tr class='tablerow'>\n"
#                 for i, cell in enumerate(row):
#                     if i >= len(headers):  # Skip cells beyond the number of headers
#                         break

#                     cell_value = cell.value if cell.value else ''
#                     # Format year values correctly
#                     if i == 0 and isinstance(cell_value, float):
#                         cell_value = int(cell_value)

#                     # Check for a hyperlink in the cell
#                     if cell.hyperlink:
#                         # Assuming the hyperlink text is at the end of the cell_value
#                         parts = str(cell_value).rsplit(' ', 1)
#                         link_text = parts[-1] if len(parts) > 1 else cell_value
#                         non_link_text = parts[0] if len(parts) > 1 else ''
#                         cell_value = f'{non_link_text} <a href="{cell.hyperlink.target}">{link_text}</a>'
#                     else:
#                         cell_value = str(cell_value)

#                     html_str += f"  <td>{cell_value}</td>\n"
#                 html_str += "</tr>\n"
#             html_str += "</table>\n\n"

#             # Write to output file
#             output_file.write(html_str)

# generate_html_from_excel_with_adjustments('./TGA_Events_Past_Champions.xlsx', 'output_html.txt')

## example CSS to use ##

# selector h3 {
#   text-align: center;
#   display: flex;x``
#   align-items: center;
#   padding: 0px;
#   margin-left: 2rem;
#   margin-right: 2rem;
#   color: white;
#   font-family: "Stag Regular", sans-serif;
# }
# .selector table {
#   border-collapse: separate;
#   border-style: none !important;
#   border-spacing: 10px;
# }
# selector td {
#   border-style : hidden!important;
#   font-size: 1rem;
# }
# .fontbold {
#   font-weight: 600;
# }
# .tablehead {
#   display: flex;
#   flex-direction: row;
#   margin-bottom: 1rem;
# }
# .tablerow {
#   background-color: #4d78c4;
#   color: white;
#   margin-bottom: 80px !important;
#   width: 100%;
#   height: 40px;
# }
# .my-button {
#   background-color: transparent;
#   border: 1px solid white;
#   color: white;
#   padding: 10px 20px;
#   font-size: 16px;
#   cursor: pointer;
#   width: fit-content;
#   font-family: "Stag Sans Bold", sans-serif;
# }

# .selector h3 {
#   font-family: "Stag bold", sans-serif;
#   font-size: 1.7rem;
#   letter-spacing: 1.5px;
# }

# @media only screen and (max-width: 600px) {
#   .selector table tr td {
#     font-size: 10px !important;
#     height: 30px !important;
#   }
#   .theme-button {
#     font-size: 12px !important;
#     margin-right: 10px !important;
#     width: 140px !important;
#   }
#   .selector h3 {
#     font-size: 16px;
#     max-width: 160px;
#   }
# }

